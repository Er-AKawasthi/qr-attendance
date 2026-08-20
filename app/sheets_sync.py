"""
Excel Export Module — Generates attendance.xlsx with all attendance data.
Format:
  Row 1: Roll No | Name | 20-Aug | 21-Aug | ...
  Row 2+: IIT2024001 | Amit | P | A | ...

New students are auto-added when they register.
New date columns are added each day attendance is taken.
"""

import os
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.database import get_db

EXCEL_FILE = Path(os.getenv("EXCEL_PATH", "attendance.xlsx"))


def _style_header(ws, col_count):
    """Apply beautiful styling to the header row."""
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2B3A67", end_color="2B3A67", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border


def _style_cell(cell, value):
    """Style a data cell — green P, red A."""
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(name="Calibri", size=11)

    if value == "P":
        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        cell.font = Font(name="Calibri", size=11, bold=True, color="006100")
    elif value == "A":
        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        cell.font = Font(name="Calibri", size=11, bold=True, color="9C0006")


def generate_excel():
    """
    Generate/update attendance.xlsx from the database.
    Returns the file path.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    with get_db() as conn:
        cursor = conn.cursor()

        # Get all students sorted by roll number
        cursor.execute("SELECT id, roll_number, name FROM students ORDER BY roll_number")
        students = [dict(row) for row in cursor.fetchall()]

        # Get all session dates (unique, sorted)
        cursor.execute("SELECT DISTINCT date FROM sessions ORDER BY date")
        dates = [row["date"] for row in cursor.fetchall()]

        if not students:
            # Empty sheet with just headers
            ws.append(["Roll No", "Name", "No data yet"])
            _style_header(ws, 3)
            wb.save(str(EXCEL_FILE))
            return str(EXCEL_FILE)

        # Format dates for column headers (20-Aug-2026 → 20-Aug)
        def fmt_date(d):
            try:
                from datetime import datetime
                dt = datetime.strptime(d, "%Y-%m-%d")
                return dt.strftime("%d-%b")
            except Exception:
                return d

        date_headers = [fmt_date(d) for d in dates]

        # Header row
        headers = ["Roll No", "Name"] + date_headers
        ws.append(headers)

        # Build attendance lookup: {student_id: {date: True}}
        cursor.execute("""
            SELECT a.student_id, s.date
            FROM attendance a
            JOIN sessions s ON a.session_id = s.id
        """)
        attendance_map = {}
        for row in cursor.fetchall():
            sid = row["student_id"]
            d = row["date"]
            if sid not in attendance_map:
                attendance_map[sid] = set()
            attendance_map[sid].add(d)

        # Data rows
        for student in students:
            row_data = [student["roll_number"], student["name"]]
            for date in dates:
                present = date in attendance_map.get(student["id"], set())
                row_data.append("P" if present else "A")
            ws.append(row_data)

        # Style everything
        _style_header(ws, len(headers))

        for row_idx in range(2, len(students) + 2):
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if col_idx > 2:
                    _style_cell(cell, cell.value)
                else:
                    cell.alignment = Alignment(horizontal="center" if col_idx == 1 else "left", vertical="center")
                    cell.font = Font(name="Calibri", size=11)
                    cell.border = Border(
                        left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"),
                    )

        # Column widths
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 22
        for col_idx in range(3, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 10

        # Freeze header row + name columns
        ws.freeze_panes = "C2"

    wb.save(str(EXCEL_FILE))
    return str(EXCEL_FILE)
